# app1.py (v50.2 - 修正權限範圍 + API雙重備援版 + 完整功能保留)

import os
import time
import requests
import json
import traceback
from datetime import datetime
from flask import Flask, request, render_template, jsonify, Response
import google.generativeai as genai
from mimetypes import guess_type
import fitz  # PyMuPDF
import io
import csv

# --- 新增：Google Drive API 相關套件 ---
from google.auth.transport.requests import Request as GoogleRequest
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# --- 新增：環境變數管理套件 ---
from dotenv import load_dotenv

# --- 新增：載入 .env 檔案 (本地開發用) ---
load_dotenv()

# --- 確認/新增函式庫 ---
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- 設定 ---
INPUT_FOLDER = "uploads"
PDF_CONVERSION_DPI = 300
app = Flask(__name__)

app.config['UPLOAD_FOLDER'] = INPUT_FOLDER
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# --- API Key 設定 ---
GEMINI_API_KEY = os.getenv('GOOGLE_API_KEY')

if not GEMINI_API_KEY:
    print("⚠️ 嚴重警告：未偵測到 GOOGLE_API_KEY！程式將無法辨識發票。")
else:
    os.environ['GOOGLE_API_KEY'] = GEMINI_API_KEY
    genai.configure(api_key=GEMINI_API_KEY)

# --- Google Drive 權限設定 (關鍵修正) ---
# 改回 drive.file 以匹配原本的 token.json，解決 invalid_scope 錯誤
SCOPES = ['https://www.googleapis.com/auth/drive.file']

# --- 統一發票字軌設定 (依年度區分) ---

# 114年 (2025)
INVOICE_PREFIX_MAP_2025 = { 
    'PT': '21', 'HT': '21', 'KT': '21', 'MT': '21', 'RT': '21', 'TT': '21',
    'HV': '22', 'HW': '22', 'HX': '22', 'HY': '22', 'KV': '22', 'KW': '22', 'KX': '22',
    'KY': '22', 'MV': '22', 'MW': '22', 'MX': '22', 'MY': '22', 'PV': '22', 'PW': '22',
    'PX': '22', 'PY': '22', 'RV': '22', 'RW': '22', 'RX': '22', 'RY': '22', 'TV': '22',
    'TW': '22', 'TX': '22', 'TY': '22', 'MW': '22'
}

# 115年 (2026) - 新增
INVOICE_PREFIX_MAP_2026 = { 
    'VT': '21', 'VU': '21', 'XV': '21', 'XW': '21', 'ZX': '21', 'ZY': '21',
    'CA': '21', 'CB': '21', 'EC': '21', 'ED': '21', 'GE': '21', 'GF': '21',
    'VV': '22', 'XX': '22', 'ZZ': '22', 'CC': '22', 'EE': '22', 'GG': '22', 
    'VW': '22', 'XY': '22', 'AA': '22', 'CD': '22', 'EF': '22', 'GH': '22', 
    'VX': '22', 'XZ': '22', 'AB': '22', 'CE': '22', 'EG': '22', 'GJ': '22', 
    'VY': '22', 'YA': '22', 'AC': '22', 'CF': '22', 'EH': '22', 'GK': '22'
}

# --- Google Drive 輔助函式 ---
def get_drive_service():
    """取得 Google Drive Service (OAuth 2.0)"""
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(GoogleRequest())
        else:
            if os.path.exists('credentials.json'):
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=0)
            else:
                print("❌ 錯誤：找不到 credentials.json，無法使用 Google Drive 功能")
                return None
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)

def download_files_from_drive_folder(folder_id):
    """從指定 Drive 資料夾下載所有圖片到 upload 資料夾"""
    service = get_drive_service()
    if not service: return []

    print(f"正在讀取雲端資料夾 ID: {folder_id} ...")
    
    # 搜尋資料夾內的檔案 (圖片與PDF)
    query = f"'{folder_id}' in parents and (mimeType contains 'image/' or mimeType = 'application/pdf') and trashed = false"
    results = service.files().list(q=query, fields="files(id, name, mimeType)").execute()
    items = results.get('files', [])

    downloaded_files = []
    if not items:
        print("資料夾內沒有檔案。")
        return []

    print(f"找到 {len(items)} 個檔案，開始下載...")

    for item in items:
        file_id = item['id']
        file_name = item['name']
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], file_name)
        
        # 下載檔案
        request = service.files().get_media(fileId=file_id)
        fh = io.FileIO(save_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        print(f"已下載: {file_name}")
        downloaded_files.append(file_name)
    
    return downloaded_files

# --- 核心函式 (Prompt 嚴禁更動) ---
def extract_data_with_gemini_vision(image_bytes: bytes, mime_type: str) -> list:
    if not GEMINI_API_KEY:
        print("[Error] 缺少 API Key，跳過辨識。")
        return []

    image_part = {"mime_type": mime_type, "data": image_bytes}
    prompt = f"""
    你是一位頂尖的台灣發票資料分析師。你的任務是從眼前的發票圖片中，精準地擷取結構化資訊。

    **最終輸出指示 (非常重要):**
    你的回覆**必須**只包含一個乾淨的 JSON 物件，其結構為 `{{ "receipts": [...] }}`。**絕對不要**在 JSON 前後或內部加入任何說明文字、分析過程、註解或 ```json ``` 標籤。

    **擷取規則:**
    1.  **[視覺辨識優先]** 直接根據圖片上的視覺資訊進行判斷，忽略微小的污漬或印刷瑕疵。特別注意帶有斜線的'0'可能被看成'8'或'6'或'9'。
    2.  **發票號碼 (invoice_number):** 格式為 "XX12345678"。
    3.  **交易日期 (date):** 擷取年月日。如果是民國年(如114)，請轉換為西元年(2025)。最終格式為 "YYYY-MM-DD"。
    4.  **交易時間 (time):** 擷取時:分:秒，格式為 "HH:MM:SS"。如果圖片上沒有，則為空字串 ""。
    5.  **賣方統一編號 (seller_vat):** 8位數字。直接從圖片中的賣方資訊區塊（通常有公司印章或明確標示）尋找。**忽略印章上的裝飾性符號（如梅花✳、星號*等），只提取數字**。
        * **主要線索：** 優先尋找有「賣方、「統編」、「統一編號」、「NO.」、「#」等明確標示的8位數字，而在收銀機三聯式發票中，這方統一編號有時會緊接在交易之後，並用#做連結。
        * **位置線索：** 在電子或手開發票中，賣方統編常常出現在「專用章」、「TEL」、「負責人」字樣附近。
        * **順序線索：** 在收銀機發票中，如果沒有明確標示，它有時是發票號碼後的第二組8位數字。
        * **修正規則：** 如果辨識出的數字串超過8位（例如 68488162716），應由左至右每8位數字為一組，逐一比對於財政部稅籍登記資料之營業人名稱(businessNm)之前兩個字，是否與本張發票有的公司名稱前兩字相同(例如68488162無稅籍登記資料，改用下一組84881627，找到「鼎祥氣體有限公司」，其前兩字與「專用章」附近出現的「7鼎祥氣體有限公司」前兩字一致，得出正確答案)。
    6.  **買方統一編號 (buyer_vat):** 尋找**除了**「賣方統一編號」之外的另一組8位數字，通常位於「買受人」字樣附近。如果沒有，則為 "N/A"。
    7.  **金額總計 (total_amount):** 發票上的含稅總金額，必須是整數。優先順序為：「總計」、「縂計」> 「合計」 > 「應收金額」，且三聯式發票通常有加總關係可以核對，例如未稅100+進項稅額5=金額總計105,。
    8.  如果任何欄位找不到，請使用 "N/A" 作為值。

    **JSON 輸出格式範例:**
    {{
      "receipts": [
        {{ "invoice_number": "MW25046739", "date": "2025-05-27", "time": "13:10", "seller_vat": "49280041", "buyer_vat": "83251000", "total_amount": 465 }}
      ]
    }}
    """
    try:
        model = genai.GenerativeModel("gemini-3-pro-preview") 
        response = model.generate_content([prompt, image_part])
        cleaned_response_text = response.text.strip()
        json_start = cleaned_response_text.find('{')
        json_end = cleaned_response_text.rfind('}')
        if json_start != -1 and json_end != -1:
            json_str = cleaned_response_text[json_start:json_end+1]
            data = json.loads(json_str)
            return data.get("receipts", [])
        else:
            print("[Gemini Vision Warning] 回應中未找到有效的 JSON 物件。")
            return []
    except Exception as e:
        print(f"[Gemini Vision Error] 解析失敗: {e}")
        return []

def is_valid_vat_number(vat: str) -> bool:
    if not vat or not vat.isdigit() or len(vat) != 8: return False
    multipliers = [1, 2, 1, 2, 1, 2, 4, 1]; total = 0
    for i in range(8): product = int(vat[i]) * multipliers[i]; total += (product // 10) + (product % 10)
    if total % 10 == 0: return True
    if vat[6] == '7' and (total + 1) % 10 == 0: return True
    return False

def correct_vat_number(vat: str) -> str:
    if is_valid_vat_number(vat): return vat
    error_indices = [i for i, char in enumerate(vat) if char in ('8', '6')]
    for i in error_indices:
        temp_vat_list = list(vat); temp_vat_list[i] = '0'
        corrected_vat = "".join(temp_vat_list)
        if is_valid_vat_number(corrected_vat): print(f"統一編號修正成功: {vat} -> {corrected_vat}"); return corrected_vat
    return vat

# --- 增強版公司查詢 (含備援) ---
def get_company_info_from_fia_api(vat_number: str) -> dict:
    if not vat_number or vat_number == 'N/A' or not vat_number.isdigit(): 
        return {"name": "N/A", "address": ""}
    
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    
    # --- 策略 A: 優先嘗試財政部官方 API ---
    try:
        api_url = f"https://eip.fia.gov.tw/OAI/api/businessRegistration/{vat_number}"
        response = requests.get(api_url, headers=headers, timeout=3)
        if response.status_code == 200:
            data = response.json()
            company_name = data.get("businessNm", "")
            company_address = data.get("businessAddress", "")
            if company_name:
                full_width_chars = "０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏpqrstuvwxyz"
                half_width_chars = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
                translation_table = str.maketrans(full_width_chars, half_width_chars)
                return {
                    "name": company_name.translate(translation_table), 
                    "address": company_address.translate(translation_table)
                }
    except Exception: pass # 失敗則安靜跳過，不報錯

    # --- 策略 B: 備援方案 - 爬取台灣公司網 ---
    try:
        print(f"啟動備援查詢: {vat_number}")
        backup_url = f"https://www.twincn.com/item.aspx?no={vat_number}"
        response = requests.get(backup_url, headers=headers, timeout=5)
        if response.status_code == 200:
            content = response.text
            start_marker = '<meta property="og:title" content="'
            end_marker = '" />'
            start_index = content.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = content.find(end_marker, start_index)
                if end_index != -1:
                    raw_title = content[start_index:end_index]
                    company_name = raw_title.split('-')[0].strip()
                    # 備援通常抓不到地址，留空
                    return {"name": company_name, "address": ""} 
    except Exception: pass

    return {"name": "查無資料(連線失敗)", "address": ""}

def enrich_and_finalize_data(raw_receipts: list, source_filename: str) -> list:
    final_receipts = []
    for raw_receipt in raw_receipts:
        try: total = int(raw_receipt.get("total_amount", 0))
        except (ValueError, TypeError): total = 0
        tax_exclusive_amount = round(total / 1.05) if total > 0 else 0; tax_amount = total - tax_exclusive_amount if total > 0 else 0
        date_part = raw_receipt.get("date", "N/A"); day_of_week = "N/A"
        
        # --- 判斷年度並選擇對應字軌表 ---
        selected_map = INVOICE_PREFIX_MAP_2025 # 預設使用 2025 (舊表)
        if date_part != "N/A":
            try: 
                dt = datetime.strptime(date_part, '%Y-%m-%d')
                weekdays = ["一", "二", "三", "四", "五", "六", "日"]
                day_of_week = weekdays[dt.weekday()]
                
                # 年度判斷邏輯
                if dt.year == 2026:
                    selected_map = INVOICE_PREFIX_MAP_2026
                # 可以在此擴充更多年份
                
            except ValueError: day_of_week = "格式錯誤"
        
        seller_vat = raw_receipt.get("seller_vat", "N/A"); buyer_vat = raw_receipt.get("buyer_vat", "N/A")
        corrected_seller_vat = correct_vat_number(seller_vat); corrected_buyer_vat = correct_vat_number(buyer_vat)
        seller_info = get_company_info_from_fia_api(corrected_seller_vat); time.sleep(0.5); buyer_info = get_company_info_from_fia_api(corrected_buyer_vat)
        invoice_number = raw_receipt.get("invoice_number", "N/A"); prefix = invoice_number[:2].upper() if invoice_number and len(invoice_number) == 10 else ""
        
        # --- 使用依年度選定的字軌表 ---
        format_code_str = selected_map.get(prefix, '25');
        
        try: format_code_int = int(format_code_str)
        except (ValueError, TypeError): format_code_int = 25
        receipt = {
            "統一發票號碼": invoice_number, "格式": format_code_int,
            "交易日期": date_part, "星期": day_of_week, "交易時間": raw_receipt.get("time", ""),
            "賣方統一編號": corrected_seller_vat, "賣方名稱": seller_info["name"], "賣方營業地址": seller_info["address"],
            "買方統一編號": corrected_buyer_vat, "買方名稱": buyer_info["name"], "買方營業地址": buyer_info["address"],
            "未稅金額": tax_exclusive_amount, "進項稅額": tax_amount, "金額總計": total, "來源檔案": source_filename,
        }
        final_receipts.append(receipt)
    return final_receipts

# --- Flask 路由 (Routes) ---
@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

# --- 新增：從 Google Drive 匯入處理 ---
@app.route('/process_drive_folder', methods=['POST'])
def process_drive_folder():
    try:
        # 從請求中獲取 folder_id，若無則使用環境變數
        json_data = request.json or {}
        folder_id = json_data.get('folder_id') or os.getenv('GDRIVE_FOLDER_ID')

        if not folder_id:
            return jsonify({"error": "未提供 Folder ID 且 .env 中也未設定"}), 400

        # 1. 下載檔案
        downloaded_files = download_files_from_drive_folder(folder_id)
        if not downloaded_files:
            return jsonify({"error": "雲端資料夾為空或下載失敗"}), 404

        all_results = []
        # 2. 處理檔案 (邏輯與 process_image 相同)
        for filename in downloaded_files:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            try:
                mime_type = guess_type(filepath)[0]; raw_receipts = []
                if mime_type in ["image/jpeg", "image/png", "image/webp"]:
                    with open(filepath, "rb") as f: image_bytes = f.read()
                    raw_receipts = extract_data_with_gemini_vision(image_bytes, mime_type)
                elif mime_type == "application/pdf":
                    doc = fitz.open(filepath)
                    for page_num, page in enumerate(doc):
                        print(f"處理 PDF '{filename}' 的第 {page_num + 1} 頁...")
                        pix = page.get_pixmap(dpi=PDF_CONVERSION_DPI); img_bytes = pix.tobytes("png")
                        page_receipts = extract_data_with_gemini_vision(img_bytes, "image/png"); raw_receipts.extend(page_receipts)
                    doc.close()
                else: 
                    # 若下載到非圖片檔，略過不報錯
                    print(f"略過非支援檔案: {filename}")
                    continue
                    
                finalized_results = enrich_and_finalize_data(raw_receipts, filename); all_results.extend(finalized_results)
            except Exception as e:
                print(f"處理檔案 {filename} 時發生錯誤: {e}"); traceback.print_exc()
                all_results.append({"來源檔案": filename, "統一發票號碼": f"處理失敗: {e}",})
            finally:
                # 處理完刪除本地暫存檔，保持乾淨
                if os.path.exists(filepath): os.remove(filepath)

        return jsonify({"results": all_results})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route('/process_image', methods=['POST'])
def process_image():
    uploaded_files = request.files.getlist('receipt_image');
    if not uploaded_files or uploaded_files[0].filename == '': return jsonify({"error": "沒有選擇任何檔案"}), 400
    all_results = []
    for file in uploaded_files:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename); file.save(filepath)
        try:
            mime_type = guess_type(filepath)[0]; raw_receipts = []
            if mime_type in ["image/jpeg", "image/png", "image/webp"]:
                with open(filepath, "rb") as f: image_bytes = f.read()
                raw_receipts = extract_data_with_gemini_vision(image_bytes, mime_type)
            elif mime_type == "application/pdf":
                doc = fitz.open(filepath)
                for page_num, page in enumerate(doc):
                    print(f"處理 PDF '{file.filename}' 的第 {page_num + 1} 頁...")
                    pix = page.get_pixmap(dpi=PDF_CONVERSION_DPI); img_bytes = pix.tobytes("png")
                    page_receipts = extract_data_with_gemini_vision(img_bytes, "image/png"); raw_receipts.extend(page_receipts)
                doc.close()
            else: raise Exception(f"不支援的檔案格式: {mime_type}")
            finalized_results = enrich_and_finalize_data(raw_receipts, file.filename); all_results.extend(finalized_results)
        except Exception as e:
            print(f"--- 處理檔案 {file.filename} 時發生嚴重錯誤 ---"); traceback.print_exc()
            all_results.append({"來源檔案": file.filename, "統一發票號碼": f"處理失敗: {e}",})
        finally:
            if os.path.exists(filepath): os.remove(filepath)
    return jsonify({"results": all_results})

@app.route('/generate_gv', methods=['POST'])
def generate_gv():
    json_data = request.json
    results = json_data.get('results', [])
    account_payable_code = json_data.get('account_payable_code', '')

    if not results:
        return jsonify({"error": "沒有資料可供下載"}), 400

    def convert_format_code_to_type(code):
        return "Q" if str(code) == "22" else "I"

    template_csv_path = "GV example.xls - PURDATA.csv"
    try:
        with open(template_csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f); header_row = next(reader)
            while len(header_row) < 24: header_row.append("")
            header_row = header_row[:24]
    except Exception as e:
        print(f"讀取範本檔表頭時發生錯誤(使用預設): {e}")
        header_row = [
            "序號", "公司別", "發票號碼", "稅籍編號", "統一編號", "記帳點",
            "發票/憑證類別", "格式代號", "單據憑證日期", 
            "傳票日期", "申報年月", "銷售人統一編號", "銷售人名稱", "課稅別", "進貨折讓區分", 
            "未稅金額", "進項稅額", "金額總計", "進項稅性質別", "扣抵代號",
            "彙總張數", "彙加註記", "應付立帳號碼", "備註" 
        ]

    data_for_excel = []; data_for_excel.append(header_row)

    for index, row_data in enumerate(results):
        transaction_date = row_data.get("交易日期", "")
        formatted_date_for_I_str = transaction_date.replace("-", "") if transaction_date else ""
        formatted_date_for_I_val = formatted_date_for_I_str if formatted_date_for_I_str else None

        format_code = row_data.get("格式", 25)
        try: format_code_int = int(format_code)
        except (ValueError, TypeError): format_code_int = 25
        format_code_val = str(format_code_int) if format_code_int is not None else None

        try: tax_exclusive = int(row_data.get("未稅金額", 0))
        except (ValueError, TypeError): tax_exclusive = 0
        try: tax = int(row_data.get("進項稅額", 0))
        except (ValueError, TypeError): tax = 0
        try: total = int(row_data.get("金額總計", 0))
        except (ValueError, TypeError): total = 0

        gv_dict = {
            "序號": index + 1,
            "憑證日期": "HD",
            "發票號碼": str(row_data.get("統一發票號碼", "")),
            "憑證類別": "721401318",
            "憑證號碼": "03251000",
            "交易幣別": 1,
            "發票/憑證類別": str(convert_format_code_to_type(format_code_int)),
            "格式代號": format_code_val,
            "單據憑證日期": formatted_date_for_I_val,
            "專案代號": None,
            "部門代號": None,
            "銷售人統一編號": str(row_data.get("賣方統一編號", "")),
            "銷售人名稱": str(row_data.get("賣方名稱", "")),
            "摘要": None,
            "": None,
            "未稅金額": tax_exclusive if tax_exclusive != 0 else None,
            "進項稅額": tax if tax != 0 else None,
            "金額總計": total if total != 0 else None,
            "結帳號碼": "126200",
            "結帳狀態": 1,
            "結帳人": 0,
            "來源碼": "N",
            "應付立帳號碼": str(account_payable_code) if account_payable_code else None,
            "傳票編號": None
        }

        gv_dict["序號"] = str(gv_dict["序號"]) if gv_dict["序號"] is not None else None
        
        current_row_list = [gv_dict.get(header_name) for header_name in header_row]
        data_for_excel.append(current_row_list)

    wb = Workbook(); ws = wb.active; ws.title = "PURDATA"

    for row_idx, row_data in enumerate(data_for_excel, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = 'General'
            if col_idx == 1 or (3 <= col_idx <= 15):
                cell.number_format = '@' 
                cell.value = str(cell_value) if cell_value is not None else '' 
            else:
                 if isinstance(cell_value, (int, float)):
                      cell.value = cell_value
                 else:
                      cell.value = str(cell_value) if cell_value is not None else ''

            cell.font = Font(name='Microsoft JhengHei', size=14)

            if row_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_idx == 1 or col_idx == 8: 
                cell.alignment = Alignment(horizontal='left', vertical='center')
            elif cell.number_format == 'General' and isinstance(cell.value, (int, float)): 
                 cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    for col_idx in range(1, len(header_row) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        header_cell = ws.cell(row=1, column=col_idx); header_len = 0
        if header_cell.value is not None:
             for char in str(header_cell.value): header_len += 2 if '\u4e00' <= char <= '\u9fff' else 1
        max_length = header_len
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value is not None:
                    cell_len = 0; value_str = str(cell.value)
                    for char in value_str: cell_len += 2 if '\u4e00' <= char <= '\u9fff' else 1
                    if cell_len > max_length: max_length = cell_len
            except: pass
        adjusted_width = (max_length + 2); ws.column_dimensions[column_letter].width = adjusted_width

    output_buffer = io.BytesIO(); wb.save(output_buffer); excel_data = output_buffer.getvalue()

    return Response(
        excel_data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=GV_output.xlsx"}
    )

if __name__ == '__main__':
    print("--- 發票批次辨識與剖析程式 (v50.2 - 修正權限 + 備援API版) ---")
    app.run(port=5000, debug=True)